"""
╔══════════════════════════════════════════════════════════════╗
║        AMEI — Automação de Encaixes (Agenda Noturna)         ║
║        AmorSaúde · Selenium + Python + Tkinter GUI           ║
╚══════════════════════════════════════════════════════════════╝

INSTALAÇÃO:
    pip install selenium webdriver-manager openpyxl

EMPACOTAR COMO EXE:
    pip install pyinstaller
    pyinstaller --onefile --windowed amei_encaixe.py

USO:
    1. Selecione o arquivo .xlsx na interface
    2. Clique em Iniciar
    3. Faça login, abra a agenda do médico, deixe os horários visíveis
    4. Clique em Continuar na interface — o script agenda tudo automaticamente

FORMATO DA PLANILHA (colunas obrigatórias, na primeira aba):
    cpf             | horario_real
    123.456.789-00  | 09:30
    987.654.321-00  | 10:00

PARA EMPACOTAR COMO EXE:
pyinstaller --onefile --windowed amei_encaixe.py

CASO QUEIRA EMPACOTAR COMO EXE E TENHA ERROS DE IMPORTAÇÃO DE TKINTER OU SELENIUM, TENTE ESTE COMANDO:  
pyinstaller --onefile --windowed --collect-all selenium --collect-all webdriver_manager --collect-all openpyxl amei_encaixe.py

"""

from os import path
import time
import os
import sys
import threading
from datetime import datetime

# ── GUI ────────────────────────────────────────────────────
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import tkinter.font as tkfont

# ── Dependências externas ──────────────────────────────────
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, ElementClickInterceptedException
    )
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.service import Service
except ImportError:
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk(); root.withdraw()
    messagebox.showerror("Dependência faltando", "Execute:\npip install selenium webdriver-manager")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk(); root.withdraw()
    messagebox.showerror("Dependência faltando", "Execute:\npip install openpyxl")
    sys.exit(1)

# ────────────────────────────────────────────
# CONFIGURAÇÕES
# ────────────────────────────────────────────
DEFAULT_URL   = "https://amei.amorsaude.com.br"
WAIT_TIMEOUT  = 10
PAUSA_CURTA   = 0.3
PAUSA_REDE    = 1.0
PAUSA_CPF     = 1.5

# ─────────────────────────────────────────────
# CORES DA INTERFACE
# ─────────────────────────────────────────────
BG_DARK    = "#0d1117"
BG_CARD    = "#161b22"
BG_INPUT   = "#1c2128"
ACCENT     = "#2ea043"
ACCENT_HOV = "#3fb950"
ACCENT2    = "#1f6feb"
TEXT_MAIN  = "#e6edf3"
TEXT_DIM   = "#7d8590"
TEXT_ERR   = "#f85149"
TEXT_WARN  = "#d29922"
TEXT_OK    = "#3fb950"
BORDER     = "#30363d"

# ─────────────────────────────────────────────
# LEITURA DO XLSX
# ─────────────────────────────────────────────
def ler_xlsx(caminho: str) -> list:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active
    cabecalho = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    try:
        idx_cpf  = cabecalho.index("cpf")
        idx_hora = cabecalho.index("horario_real")
    except ValueError:
        raise ValueError(f"Colunas 'cpf' e/ou 'horario_real' não encontradas.\nCabeçalho: {cabecalho}")

    registros = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cpf  = str(row[idx_cpf]).strip()  if row[idx_cpf]  is not None else ""
        hora = str(row[idx_hora]).strip() if row[idx_hora] is not None else ""
        if not cpf or cpf.lower() == "none":
            continue
        if not hora or hora.lower() == "none":
            continue
        if len(hora) > 5 and hora[2] == ":":
            hora = hora[:5]
        registros.append({"cpf": cpf, "horario_real": hora})

    if not registros:
        raise ValueError("Nenhum registro válido na planilha.")
    return registros

# ───────────────────────────────────────────
# BROWSER
# ───────────────────────────────────────────
def iniciar_browser(url: str) -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts
    )
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    driver.get(url)
    return driver

# ─────────────────────────────────────────────
# HELPERS SELENIUM
# ─────────────────────────────────────────────
def aguardar_sem_overlay(driver, timeout=5):
    for sel in ["div[style*='rgba(255, 255, 255, 0.5)']", "div[class*='loading']"]:
        try:
            WebDriverWait(driver, timeout).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, sel))
            )
        except TimeoutException:
            pass

def clicar_seguro(driver, elemento, tentativas=3):
    for t in range(tentativas):
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elemento)
            aguardar_sem_overlay(driver)
            elemento.click()
            return True
        except ElementClickInterceptedException:
            if t < tentativas - 1:
                time.sleep(0.5)
            else:
                driver.execute_script("arguments[0].click();", elemento)
                return True
        except Exception as e:
            if t == tentativas - 1:
                raise e
            time.sleep(0.5)
    return False

def buscar_horarios_livres(driver) -> list:
    todos = []
    ultima_altura = 0

    while True:
        # Busca botões de horário livre visíveis
        candidatos = [b for b in driver.find_elements(By.CSS_SELECTOR,
            "button.action-button.action-create, button[class*='action-create']")
            if b.is_displayed()]

        if not candidatos:
            for bloco in driver.find_elements(By.CSS_SELECTOR, "div.cal-event"):
                if not bloco.is_displayed():
                    continue
                try:
                    texto = bloco.find_element(
                        By.CSS_SELECTOR, "p.patient-title-name").text.strip()
                    if len(texto) == 5 and texto[2] == ":" and texto.replace(":", "").isdigit():
                        candidatos.append(bloco)
                except NoSuchElementException:
                    continue

        # Adiciona novos (evita duplicatas por posição Y)
        posicoes_existentes = {e.location["y"] for e in todos}
        for c in candidatos:
            if c.location["y"] not in posicoes_existentes:
                todos.append(c)
                posicoes_existentes.add(c.location["y"])

        # Rola para baixo
        nova_altura = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.8)

        nova_altura2 = driver.execute_script("return document.body.scrollHeight")
        if nova_altura2 == ultima_altura:
            break  # chegou ao fim da página
        ultima_altura = nova_altura2

    todos.sort(key=lambda e: e.location["y"])
    return todos

# ─────────────────────────────────────────────
# PASSOS DO AGENDAMENTO
# ─────────────────────────────────────────────
def preencher_cpf_e_obs(driver, cpf: str, horario_real: str) -> bool:
    aguardar_sem_overlay(driver)
    try:
        campo_cpf = WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.ID, "cpf"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo_cpf)
        aguardar_sem_overlay(driver)
        driver.execute_script("arguments[0].click();", campo_cpf)
        for digito in cpf.replace(".", "").replace("-", "").replace(" ", ""):
            campo_cpf.send_keys(digito)
            time.sleep(0.04)

        # Aguarda um pouco para o sistema processar o CPF
        time.sleep(PAUSA_CPF)

        # Verifica se apareceu modal "Paciente não encontrado"
        try:
            modal_cancelar = WebDriverWait(driver, 4).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-cancel"))
            )
            # Modal apareceu — clica em Cancelar
            clicar_seguro(driver, modal_cancelar)
            time.sleep(0.5)

            # Clica no Cancelar do formulário para fechar a tela de agendamento
            for b in driver.find_elements(By.CSS_SELECTOR,
                "button.secondary-button.action-button, button[class*='secondary-button']"):
                if "cancelar" in b.text.strip().lower() and b.is_displayed():
                    clicar_seguro(driver, b)
                    break
            time.sleep(0.5)

            raise Exception("Paciente não cadastrado no sistema")

        except TimeoutException:
            # Modal não apareceu — paciente encontrado, segue normalmente
            pass

        # Aguarda nome aparecer
        try:
            WebDriverWait(driver, PAUSA_CPF * 3).until(
                lambda d: (d.find_element(By.CSS_SELECTOR,
                    "input[formcontrolname='name'], input[id='name']"
                ).get_attribute("value") or "").strip() != ""
            )
        except (TimeoutException, NoSuchElementException):
            time.sleep(PAUSA_CPF)

        # Observação
        campo_obs = None
        for sel in [(By.ID, "observation"),
                    (By.CSS_SELECTOR, "textarea[formcontrolname='observation']")]:
            try:
                campo_obs = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(sel))
                break
            except TimeoutException:
                continue
        if campo_obs:
            driver.execute_script("arguments[0].click();", campo_obs)
            campo_obs.clear()
            campo_obs.send_keys(f"{horario_real}") #Caso queira escrever algo a mais na observação, é só modificar esta linha
        return True

    except Exception as e:
        raise Exception(f"{e}")


def incluir_procedimento(driver) -> bool:
    try:
        clicar_seguro(driver, WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.add-procedure-button"))
        ))
        time.sleep(PAUSA_CURTA)
        clicar_seguro(driver, WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.element_to_be_clickable((By.ID, "AdProcedimento"))
        ))
        time.sleep(PAUSA_CURTA)
        opcoes = [o for o in WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "mat-option, .mat-option"))
        ) if o.is_displayed()]
        if not opcoes:
            raise Exception("Sem opções no dropdown")
        clinica, areas, primeira = None, None, None
        for o in opcoes:
            t = o.text.strip().lower()
            if primeira is None:
                primeira = o
            if "clínica médica" in t or "clinica medica" in t:
                clinica = o
            if "áreas de atuação" in t or "areas de atuacao" in t:
                areas = o

        escolha = clinica or areas or primeira
        clicar_seguro(driver, escolha)
        time.sleep(PAUSA_CURTA)
        clicar_seguro(driver, WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.element_to_be_clickable((By.XPATH,
                "//button[contains(@class,'mat-flat-button') and .//span[contains(text(),'Adicionar')]]"
            ))
        ))
        time.sleep(PAUSA_CURTA)
        return True
    except Exception as e:
        raise Exception(f"Falha procedimento: {e}")


def confirmar_agendamento(driver) -> bool:
    try:
        clicar_seguro(driver, WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.element_to_be_clickable((By.XPATH,
                "//button[.//span[contains(text(),'Confirmar')]]"
            ))
        ))
        return True
    except Exception:
        for b in driver.find_elements(By.TAG_NAME, "button"):
            if "confirmar" in b.text.lower() and b.is_displayed():
                clicar_seguro(driver, b)
                return True
        raise Exception("Botão Confirmar não encontrado")


def fechar_modal_sucesso(driver) -> bool:
    try:
        clicar_seguro(driver, WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm"))
        ))
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".swal2-container"))
        )
        time.sleep(PAUSA_REDE)
        return True
    except Exception:
        for b in driver.find_elements(By.TAG_NAME, "button"):
            if b.text.strip().lower() == "ok" and b.is_displayed():
                clicar_seguro(driver, b)
                time.sleep(PAUSA_REDE)
                return True
        raise Exception("Modal de sucesso não encontrado")

# ─────────────────────────────────────────────
# LOG XLSX
# ─────────────────────────────────────────────
def salvar_log(resultados: list, pasta: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ws.append(["CPF", "Horário Real", "Status"])
    for r in resultados:
        ws.append([r["cpf"], r["horario_real"], r["status"]])
    nome = os.path.join(pasta, f"resultado_encaixes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(nome)
    return nome

# ─────────────────────────────────────────────
# INTERFACE GRÁFICA
# ─────────────────────────────────────────────
class AmeiApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AMEI — Automação de Encaixes")
        self.root.geometry("780x620")
        self.root.resizable(False, False)
        self.root.configure(bg=BG_DARK)

        self.xlsx_path   = tk.StringVar()
        self.url_var     = tk.StringVar(value=DEFAULT_URL)
        self.driver      = None
        self.running     = False
        self.wait_event  = threading.Event()

        self._build_ui()

    # ── UI ──────────────────────────────────
    def _build_ui(self):
        # Título
        header = tk.Frame(self.root, bg=BG_DARK)
        header.pack(fill="x", padx=24, pady=(24, 0))

        tk.Label(header, text="AMEI", font=("Courier New", 28, "bold"),
                 bg=BG_DARK, fg=ACCENT).pack(side="left")
        tk.Label(header, text="  Automação de Encaixes · AmorSaúde",
                 font=("Courier New", 11), bg=BG_DARK, fg=TEXT_DIM).pack(side="left", pady=(10,0))

        # Separador
        tk.Frame(self.root, bg=BORDER, height=1).pack(fill="x", padx=24, pady=12)

        # Card de configuração
        card = tk.Frame(self.root, bg=BG_CARD, relief="flat", bd=0)
        card.pack(fill="x", padx=24, pady=(0, 12))

        inner = tk.Frame(card, bg=BG_CARD)
        inner.pack(fill="x", padx=16, pady=14)

        # Linha XLSX
        tk.Label(inner, text="Planilha (.xlsx)", font=("Courier New", 10, "bold"),
                 bg=BG_CARD, fg=TEXT_DIM).grid(row=0, column=0, sticky="w", pady=(0,6))

        xlsx_row = tk.Frame(inner, bg=BG_CARD)
        xlsx_row.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        inner.columnconfigure(0, weight=1)

        self.xlsx_entry = tk.Entry(xlsx_row, textvariable=self.xlsx_path,
                                   font=("Courier New", 10), bg=BG_INPUT, fg=TEXT_MAIN,
                                   relief="flat", bd=0, insertbackground=TEXT_MAIN,
                                   highlightthickness=1, highlightbackground=BORDER,
                                   highlightcolor=ACCENT2)
        self.xlsx_entry.pack(side="left", fill="x", expand=True, ipady=6, ipadx=8)

        btn_browse = tk.Button(xlsx_row, text="  Selecionar  ",
                               font=("Courier New", 10, "bold"),
                               bg=ACCENT2, fg=TEXT_MAIN, relief="flat", bd=0,
                               activebackground="#388bfd", activeforeground=TEXT_MAIN,
                               cursor="hand2", command=self._selecionar_xlsx)
        btn_browse.pack(side="left", padx=(8, 0), ipady=6)

        btn_modelo = tk.Button(xlsx_row, text="  Baixar modelo  ",
                       font=("Courier New", 10, "bold"),
                       bg=BG_CARD, fg=TEXT_DIM, relief="flat", bd=0,
                       activebackground=BORDER, activeforeground=TEXT_MAIN,
                       cursor="hand2", command=self._baixar_modelo)
        btn_modelo.pack(side="left", padx=(8, 0), ipady=6)
        
        # Linha URL
        tk.Label(inner, text="URL do sistema", font=("Courier New", 10, "bold"),
                 bg=BG_CARD, fg=TEXT_DIM).grid(row=2, column=0, sticky="w", pady=(0,6))

        self.url_entry = tk.Entry(inner, textvariable=self.url_var,
                                  font=("Courier New", 10), bg=BG_INPUT, fg=TEXT_MAIN,
                                  relief="flat", bd=0, insertbackground=TEXT_MAIN,
                                  highlightthickness=1, highlightbackground=BORDER,
                                  highlightcolor=ACCENT2)
        self.url_entry.grid(row=3, column=0, sticky="ew", ipady=6, ipadx=8)

        # Separador
        tk.Frame(self.root, bg=BORDER, height=1).pack(fill="x", padx=24, pady=(0, 12))

        # Botões de ação
        btn_frame = tk.Frame(self.root, bg=BG_DARK)
        btn_frame.pack(fill="x", padx=24, pady=(0, 12))

        self.btn_iniciar = tk.Button(btn_frame, text="▶  Iniciar",
                                     font=("Courier New", 11, "bold"),
                                     bg=ACCENT, fg="#000000", relief="flat", bd=0,
                                     activebackground=ACCENT_HOV, activeforeground="#000000",
                                     cursor="hand2", command=self._iniciar,
                                     width=16, pady=8)
        self.btn_iniciar.pack(side="left")

        self.btn_continuar = tk.Button(btn_frame, text="⏎  Continuar",
                                       font=("Courier New", 11, "bold"),
                                       bg=BG_CARD, fg=TEXT_DIM, relief="flat", bd=0,
                                       activebackground=BORDER, activeforeground=TEXT_MAIN,
                                       cursor="hand2", command=self._continuar,
                                       width=16, pady=8, state="disabled")
        self.btn_continuar.pack(side="left", padx=(10, 0))

        self.btn_parar = tk.Button(btn_frame, text="■  Parar",
                                   font=("Courier New", 11, "bold"),
                                   bg=BG_CARD, fg=TEXT_ERR, relief="flat", bd=0,
                                   activebackground=BORDER, activeforeground=TEXT_ERR,
                                   cursor="hand2", command=self._parar,
                                   width=16, pady=8, state="disabled")
        self.btn_parar.pack(side="left", padx=(10, 0))

        # Log
        log_label = tk.Frame(self.root, bg=BG_DARK)
        log_label.pack(fill="x", padx=24)
        tk.Label(log_label, text="Log de execução", font=("Courier New", 10, "bold"),
                 bg=BG_DARK, fg=TEXT_DIM).pack(side="left")
        self.lbl_status = tk.Label(log_label, text="", font=("Courier New", 10),
                                   bg=BG_DARK, fg=TEXT_DIM)
        self.lbl_status.pack(side="right")

        self.log_box = scrolledtext.ScrolledText(
            self.root, font=("Courier New", 9), bg=BG_CARD, fg=TEXT_MAIN,
            relief="flat", bd=0, state="disabled", wrap="word",
            highlightthickness=1, highlightbackground=BORDER,
            insertbackground=TEXT_MAIN
        )
        # Rodapé
        footer = tk.Frame(self.root, bg=BG_DARK)
        footer.pack(fill="x", padx=24, pady=(0, 8))
        tk.Label(footer, text="Desenvolvido por: Fernando Andrade  |  Github: https://github.com/catwithbot",
                font=("Courier New", 8), bg=BG_DARK, fg=TEXT_DIM).pack(side="left")
        self.log_box.pack(fill="both", expand=True, padx=24, pady=(4, 20))

        # Tags de cor no log
        self.log_box.tag_config("ok",   foreground=TEXT_OK)
        self.log_box.tag_config("err",  foreground=TEXT_ERR)
        self.log_box.tag_config("warn", foreground=TEXT_WARN)
        self.log_box.tag_config("info", foreground=ACCENT2)
        self.log_box.tag_config("bold", foreground=TEXT_MAIN, font=("Courier New", 9, "bold"))
    # ── HELPERS UI ──────────────────────────
    def _log(self, msg: str, tag: str = ""):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{ts}] {msg}\n", tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _set_status(self, msg: str, cor: str = TEXT_DIM):
        self.lbl_status.configure(text=msg, fg=cor)

    def _selecionar_xlsx(self):
        path = filedialog.askopenfilename(
            title="Selecione a planilha",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if path:
            self.xlsx_path.set(path)
            self._log(f"Planilha selecionada: {os.path.basename(path)}", "info")

    def _baixar_modelo(self):
        path = filedialog.asksaveasfilename(
        title="Salvar modelo",
        defaultextension=".xlsx",
        initialfile="encaixes_modelo.xlsx",
        filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Encaixes"
        ws.append(["cpf", "horario_real"])
        ws.append(["123.456.789-00", "09:30"])
        wb.save(path)
        self._log(f"Modelo salvo em: {os.path.basename(path)}", "ok")

    def _habilitar_botoes(self, iniciando=False):
        if iniciando:
            self.btn_iniciar.configure(state="disabled", bg=BG_CARD, fg=TEXT_DIM)
            self.btn_continuar.configure(state="disabled", bg=BG_CARD, fg=TEXT_DIM)
            self.btn_parar.configure(state="normal", fg=TEXT_ERR)
        else:
            self.btn_iniciar.configure(state="normal", bg=ACCENT, fg="#000000")
            self.btn_continuar.configure(state="disabled", bg=BG_CARD, fg=TEXT_DIM)
            self.btn_parar.configure(state="disabled", fg=TEXT_DIM)

    def _aguardar_usuario(self, prompt: str):
        self._log(f"⏸  {prompt}", "warn")
        self._set_status("Aguardando CONTINUAR...", TEXT_WARN)
        self.btn_continuar.configure(state="normal", bg=ACCENT2, fg=TEXT_MAIN)
        self.wait_event.clear()
        self.wait_event.wait()
        self.btn_continuar.configure(state="disabled", bg=BG_CARD, fg=TEXT_DIM)

    # ── AÇÕES ───────────────────────────────
    def _continuar(self):
        self.wait_event.set()

    def _parar(self):
        self.running = False
        self.wait_event.set()
        self._log("Execução interrompida pelo usuário.", "err")
        self._set_status("Parado.", TEXT_ERR)

    def _iniciar(self):
        if not self.xlsx_path.get():
            messagebox.showwarning("Atenção", "Selecione uma planilha .xlsx primeiro.")
            return
        if not os.path.exists(self.xlsx_path.get()):
            messagebox.showerror("Erro", "Arquivo não encontrado.")
            return
        self.running = True
        self._habilitar_botoes(iniciando=True)
        t = threading.Thread(target=self._executar, daemon=True)
        t.start()

    # ── EXECUÇÃO ────────────────────────────
    def _executar(self):
        resultados = []
        pasta_log  = os.path.dirname(self.xlsx_path.get()) or "."

        try:
            # Lê planilha
            self._log("Lendo planilha...", "info")
            try:
                registros = ler_xlsx(self.xlsx_path.get())
            except Exception as e:
                self._log(f"Erro ao ler planilha: {e}", "err")
                self._habilitar_botoes()
                return

            self._log(f"{len(registros)} paciente(s) carregado(s).", "ok")
            for r in registros:
                self._log(f"  · {r['cpf']} → {r['horario_real']}", "info")

            # Abre Chrome
            self._log("Iniciando Chrome...", "info")
            try:
                self.driver = iniciar_browser(self.url_var.get())
            except Exception as e:
                self._log(f"Erro ao abrir Chrome: {e}", "err")
                self._habilitar_botoes()
                return
            self._log("Chrome aberto.", "ok")

            # Aguarda login
            self._aguardar_usuario(
                "Faça login, abra a agenda e deixe os horários visíveis. Clique em CONTINUAR."
            )
            if not self.running:
                return

            self._set_status("Executando...", TEXT_OK)

            # Loop principal
            for i, reg in enumerate(registros, 1):
                if not self.running:
                    break

                self._log(f"─── {i}/{len(registros)} — {reg['cpf']} ───", "bold")
                status = "ERRO"

                try:
                    horarios = buscar_horarios_livres(self.driver)
                    if not horarios:
                        raise Exception("Nenhum horário livre visível na tela")

                    clicar_seguro(self.driver, horarios[0])
                    time.sleep(PAUSA_REDE)

                    preencher_cpf_e_obs(self.driver, reg["cpf"], reg["horario_real"])
                    self._log("CPF e observação preenchidos.", "ok")

                    incluir_procedimento(self.driver)
                    self._log("Procedimento incluído.", "ok")

                    confirmar_agendamento(self.driver)
                    self._log("Confirmado.", "ok")

                    fechar_modal_sucesso(self.driver)
                    self._log("Modal fechado.", "ok")

                    status = "OK"
                    self._log(f"✅ {reg['cpf']} agendado com sucesso!", "ok")

                except Exception as e:
                    status = f"ERRO: {e}"
                    self._log(f"❌ {reg['cpf']} — {e}", "err")
                    self._log("Continuando para o próximo...", "warn")

                resultados.append({
                    "cpf": reg["cpf"],
                    "horario_real": reg["horario_real"],
                    "status": status,
                })

        finally:
            # Resumo
            if resultados:
                ok_c  = sum(1 for r in resultados if r["status"] == "OK")
                err_c = len(resultados) - ok_c
                self._log("─── RESULTADO FINAL ───", "bold")
                self._log(f"✅ Sucesso: {ok_c}   ❌ Falhas: {err_c}", "ok" if err_c == 0 else "warn")

                try:
                    nome_log = salvar_log(resultados, pasta_log)
                    self._log(f"Log salvo: {os.path.basename(nome_log)}", "ok")
                except Exception as e:
                    self._log(f"Erro ao salvar log: {e}", "err")

            self._habilitar_botoes()
            self._set_status("Concluído." if self.running else "Parado.", TEXT_OK)
            self.running = False

            if self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass
                self.driver = None

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    
    from datetime import date #expiração da versão
    EXPIRA = date(2029, 12, 29)  # ← data de expiração
    if date.today() > EXPIRA:
        root = tk.Tk()
        root.withdraw()
        messagebox.showwarning(
            "Versão expirada",
            "Esta versão expirou.\nAguarde a próxima atualização.\n\nContato: github.com/catwithbot"
        )
        return
    
    root = tk.Tk()
    app = AmeiApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
